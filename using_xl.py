import openpyxl

# rows=위아래
# column=양옆

#주소 변수에 저장
home="/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/test.xlsx"

# 엑셀 불러오기
wb= openpyxl.load_workbook(home)

# 엑셀 시트 선택
ws=wb['Sheet1']

#셀 값 불러오는 법 1
print(ws.cell(row=3, column=3).value)

#셀 값 불러오는 법 2
print(ws['A1'].value)

#데이터 수정하기

oi=0
wi=0
#row 기준으로 출력
for rows in ws.iter_rows():
    for cell in rows:
        print(cell.value, end="\t\t\t")
        wi+=1
    print("")
    oi+=1
print(wi,oi)

col=[]
row=[]

# for rows in ws.iter_rows():
#     for cell in rows:
#         col.append(cell.value)
        # print(col[len(ws.iter_rows)][rows])
#
# print(rows)

# for i in range(8):
#     for j in range(8):
#         print(row[i][j],end=" ")
#     print("")

#rows=7
#
for i in range(1,41):
    for j in range(1,8):
        row.append(ws.cell(row=i, column=j).value)
    col.append(row);
    row=[]

for i in range(40):
    for j in range(7):
        print(col[i][j], end=" ")
    print("")

# # countt=0
#
# ii=0
# jj=0
# for rows in ws.iter_rows():
#     for cell in rows:
#         if(jj==7):
#             col[ii][jj]=cell.value
#             jj+=1
#         ii+=1


        # countt+=1
        # i+=1
        # print(cell.value, end="\t\t")
        # row[i][j]=cell.value
        # print(row[i][j])
    # print("")

# for i in range(40):
#     for j in range(7):
#         print(row[i][j], end=" ")
#     print(" ")

##F1~F40까지를 0으로 수정
# test="F"
# ws['F1']="수량"
# for i in range (2,41):
#     test2=test+str(i)
#     ws[test2]=0

#ws['A3']= "X"
# ws['A1']= "XXXXXXX"
# ws['C3']="테스트 숴정"
# ws['D3']= "X"
# ws['E3']= "XXX"
# ws['F3']= ""
# ws['G3']= "X"

#엑셀 저장
# wb.save(home)