from tkinter import * # tkinter의 모든 함수 가져오기
from tkinter import messagebox
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import tkinter.ttk


#from datetime improt datetime

#def time():

def edit():
    selected_item = treeview.selection()[0]
    treeview.item(selected_item,  values=("foo", "bar"))
def delete():
    selected_item = treeview.selection()[0] ## get selected item
    treeview.delete(selected_item)
def checker():
    # selected_item = treeview.selection()[0] ## get selected item
    # selected_item = treeview.selection()[0]
    # treeview.item(selected_item, values=("foo", "bar"))
    # treeview.delete(selected_item)

    # for i in treeview.get_children():
    #     treeview.delete(i)
    # win.update()
    oglist=[]

    treeview.delete(*treeview.get_children()) #내부 기록 지우기

    call_tree()
    win.update()
    # selected_item = treeview.selection()[0]
    # messagebox.showinfo("",selected_item)

# def defaultset():
#     ws = wb_data['Sheet1'] # 값만 받기, 원본 파일 사용
#     one_line=""
#
#     for i in range(1,(get_rows()+1)): #원본 파일을 리스트 박스에 출력
#         for j in range(1, 8):
#             if (str(ws.cell(row=i, column=j).value) == "None"):  # DK G의 함수를 None -> 0으로 받기
#                 one_line += "0"
#             else:
#                 one_line += str(ws.cell(row=i, column=j).value) + '  '
#
#         리스트.insert((i-1), one_line)
#         one_line=""
#
#     #임시로 입력을 받기 위해 수정한 텍스트 박스
#     빈소기간1.insert(0,"물품명")
#     빈소기간2.insert(0,"단가")
#     안치기간1.insert(0,"단위")
#     안치기간2.insert(0,"수량")

# def myFunc(): #새 파일과 시트 생성 -> 빈소에 들어간 숫자에 따라 사용되는 파일이 다름 -> 원본 시트의 목록 삭제 -> 새로운 시트의 목록 출력
#     nwb = openpyxl.Workbook() #엑셀 생성
#     pws = nwb.create_sheet("personal_info") #+sheet 이름
#     iws = nwb.create_sheet("items") #+sheet 이름 2
#
#
#     # sheet 1(personal_info)에 들어갈 정보
#     # A:ID B:고인명 C:상주명 D:빈소
#     pws['A1'] = ID.get()
#     pws['B1'] = 고인명.get()
#     pws['C1'] = 상주명.get()
#     pws['D1'] = 빈소.get()
#     room=빈소.get()
#
#
#     # sheet 2(items)에 들어갈 정보
#     #A:번호(용도 모름) B:물품코드 C:뭂품명 D:단위 E:단가 F:수량 G:금액
#     iws['A1'] = "번호"
#     iws['B1'] = "물품코드"  # 물품명 string
#     iws['C1'] = '물품명'
#     iws['D1'] = '단위'
#     iws['E1'] = '단가'
#     iws['F1'] = '수량'
#     iws['G1'] = '금액'
#
#     #원본 파일을 새로운 파일에 복사
#     for i in range(1,(get_rows()+1)):
#         for j in range(1, 8):
#             iws.cell(row=i,column=j).value=oglist[i-1][j-1]
#
#     #이름이 같으면 덮어씀
    # 빈소에 넣은 숫자에 따라 사용하는 엑셀이 달라짐

#     리스트.delete(0, get_rows()) #출력된 원본 시트 목록 제삭
#
#
#     one_line = ""
#     # 원본파일에서 복사된 새 시트의 목록들 출력
#     for i in range(1, (get_rows() + 1)):
#         for j in range(1, 8):
#             if (str(iws.cell(row=i, column=j).value) == "None"):  # DK G의 함수를 None -> 0으로 받기
#                 one_line += "0"
#             else:
#                 one_line += str(iws.cell(row=i, column=j).value)+"     \t"
#
#         리스트.insert((i - 1), one_line)
#         one_line = ""
#         리스트.insert(i,iws)
def create_room():
    nwb = openpyxl.Workbook()  # 엑셀 생성
    pws = nwb.create_sheet("info")  # +sheet 이름
    iws = nwb.create_sheet("items")  # +sheet 이름 2

    # sheet 1(personal_info)에 들어갈 정보
    # A:ID B:고인명 C:상주명 D:빈소
    pws['A1'] = ID.get()
    pws['B1'] = 고인명.get()
    pws['C1'] = 상주명.get()
    pws['D1'] = 빈소.get()
    room = 빈소.get()

    # sheet 2(items)에 들어갈 정보
    #A:번호(용도 모름) B:물품코드 C:뭂품명 D:단위 E:단가 F:수량 G:금액
    for i in range(1, (get_rows() +1)):
        for j in range(1, 8):
            iws.cell(row=i, column=j).value = oglist[i - 1][j - 1]


    # 빈소에 넣은 숫자에 따라 사용하는 엑셀이 달라짐
    if (room == "1"):
        nwb.save('/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/room_one.xlsx')
    elif (room == "2"):
        nwb.save('/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/room_two.xlsx')
    elif (room == "3"):
        nwb.save('/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/room_three.xlsx')
    elif (room == "4"):
        nwb.save('/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/room_four.xlsx')
    elif (room == "5"):
        nwb.save('/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/room_five.xlsx')
    elif (room == "6"):
        nwb.save('/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/room_six.xlsx')

    tree_maker()
def get_rows(): #원본 시트의 rows 길이를 구한다(아이템 숫자+첫 목록)
    count=0
    for rows in ws.iter_rows():
        count+=1
    return count
def get_cells(): #원본 시트의 총 셀 수를 가진다.
    count=0
    for rows in ws.iter_rows():
        for cell in rows:
            count+=1
    return count
def in_list(): #2차원 리스트에 저장 --oglist(원본 손상 없이 그대로 유지)
    row=[]
    #원본 시트 사용
    for i in range(1,(get_rows()+1)):
        for j in range(1, 8):
            row.append(ws_data.cell(row=i, column=j).value)
        oglist.append(row)
        row = []

def tree_maker(): #프린트를 위해 첫번째 row랑 column 제거
    get = []
    for i in range(1, get_rows()):
        for j in range(1, 7):
            get.append(oglist[i][j])
        treelist.append(get)
        get = []
    for i in range(len(treelist)):
        treeview.insert('', 'end', text=i + 2, values=treelist[i])
def call_tree():

    global treeview


    treelist=[]

    treeview= tkinter.ttk.Treeview(win, columns=["one", "two", "three", "four", "five", "six"],
                                    displaycolumns=["one", "two", "three", "four", "five", "six"], height=25)

    treeview.column("#0", width=40, anchor="center")
    treeview.heading("#0", text="번호", anchor="center")

    treeview.column("#1", width=100, anchor="center")
    treeview.heading("#1", text="물품코드", anchor="center")

    treeview.column("#2", width=100, anchor="center")
    treeview.heading("#2", text="물품명", anchor="center")

    treeview.column("#3", width=100, anchor="center")
    treeview.heading("#3", text="단위", anchor="center")

    treeview.column("#4", width=100, anchor="center")
    treeview.heading("#4", text="단가", anchor="center")

    treeview.column("#5", width=100, anchor="center")
    treeview.heading("#5", text="수량", anchor="center")

    treeview.column("#6", width=100, anchor="center")
    treeview.heading("#6", text="금액", anchor="center")



    get = []
    for i in range(1, get_rows()):
        for j in range(1, 7):
            get.append(oglist[i][j])
        treelist.append(get)
        get = []

    for i in range(len(treelist)):
        treeview.insert('', 'end', text=i + 2, values=treelist[i])




def close():
    win.quit()
    win.destroy()

#########################   global variable   ##########################

home = "/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/test.xlsx" #기본 물품 엑셀 위치 저장
wb_data = openpyxl.load_workbook(home, data_only=True) #값으로
ws_data=wb_data['Sheet1'] #사용 시트 지정


wb = openpyxl.load_workbook(home) #함수 그대로
ws = wb['Sheet1'] #사용 시트 지정
oglist=[]   #2차원 리스트에 값 저장할 때 사용 ->in_list()
treelist=[] #테이블에 사용되는 2차원 배열 ->tree_maker
in_list()

#Tkinter 윈도우 화면
win = Tk() # 창 생성
win.geometry("1000x720") # 창의 크기
win.title("장례식장 재고관리 프로그램 Ver1.221123") # 창의 제목
win.option_add("*Font", "맑은고딕 11") # 전체 폰트
#win.resizable(False, False) #윈도우 사이즈 조절 불가


#########################   excel   ##########################



#########################    menu   ##########################

menu = Menu(win)

menu_1 = Menu(menu, tearoff = 0)
menu_1.add_command(label = "로그인")
menu_1.add_command(label = "인쇄")
menu_1.add_command(label = "장부")
menu_1.add_separator()
menu_1.add_command(label = "종료", command = close)
menu.add_cascade(label = "메뉴", menu = menu_1)

win.config(menu = menu)

#########################   config  ##########################

#레이블 정의
ID_lab = Label(win)
ID_lab.config(text = "ID", width=10, relief="solid")
고인명_lab = Label(win)
고인명_lab.config(text = "고인명",width=10, relief="solid")
상주명_lab = Label(win)
상주명_lab.config(text = "상주명",width=10, relief="solid")
빈소_lab = Label(win)
빈소_lab.config(text = "빈소", width=10, relief="solid")
물결1_lab = Label(win)
물결1_lab.config(text = "~", width=10)
물결2_lab = Label(win)
물결2_lab.config(text = "~", width=10)
###
수납금액_lab = Label(win)
수납금액_lab.config(text = "수납금액", width=10, relief="solid")
받은금액_lab = Label(win)
받은금액_lab.config(text = "받은금액", width=10, relief="solid")
거스름돈_lab = Label(win)
거스름돈_lab.config(text = "거스름돈", width=10, relief="solid")

#엔트리 정의
ID = Entry(win)
ID.config(width=10,relief="solid",borderwidth=2)
고인명 = Entry(win)
고인명.config(width=10,relief="solid",borderwidth=2)
상주명 = Entry(win)
상주명.config(width=10,relief="solid",borderwidth=2)
빈소 = Entry(win)
빈소.config(width=60,relief="solid",borderwidth=2)
빈소기간1 = Entry(win)
빈소기간1.config(width=10,relief="solid",borderwidth=2)
안치기간1 = Entry(win)
안치기간1.config(width=10,relief="solid",borderwidth=2)
빈소기간2 = Entry(win)
빈소기간2.config(width=10,relief="solid",borderwidth=2)
안치기간2 = Entry(win)
안치기간2.config(width=10,relief="solid",borderwidth=2)
수납금액 = Entry(win)
수납금액.config(width=20,relief="solid",borderwidth=2)
받은금액 = Entry(win)
받은금액.config(width=20,relief="solid",borderwidth=2)
거스름돈 = Entry(win)
거스름돈.config(width=20,relief="solid",borderwidth=2)

#버튼 정의
저장 = Button(win, text = "저장", command=create_room) #DK command로 버튼 클릭시 def create_room() 실행
저장.config(width=10,height=2)
#btn.config(command=ID_a)
현금수납 = Button(win, text = "현금수납")
현금수납.config(width=10,height=3)
닫기 = Button(win, text = "닫기")
닫기.config(width=10,height=3,command =close)
식당판매 = Button(win, text = "edit")
식당판매.config(width=10,height=3, command=edit)
매점판매 = Button(win, text = "delete", command=delete)
매점판매.config(width=10,height=3)
Set = Button(win, text = "checker",command=checker)
Set.config(width=10,height=3)




#########################   treeview  ##########################

# treeview = tkinter.ttk.Treeview(win, columns=["one", "two","three","four","five","six"],
#                                 displaycolumns=["one", "two","three","four","five","six"],height=25)
#

call_tree()
# tree_maker()


#########################   place  ##########################

#레이블 위치
treeview.place(x=10, y=210)
ID_lab.place(x=10,y=10)
고인명_lab.place(x=210,y=10)
상주명_lab.place(x=410,y=10)
빈소_lab.place(x=10,y=50)
물결1_lab.place(x=250,y=90)
물결2_lab.place(x=250,y=130)
###
수납금액_lab.place(x=620,y=10)
받은금액_lab.place(x=620,y=60)
거스름돈_lab.place(x=620,y=110)

#엔트리 위치
ID.place(x=110,y=10)
고인명.place(x=310,y=10)
상주명.place(x=510,y=10)
빈소.place(x=110,y=50)
빈소기간1.place(x=10,y=180)
안치기간1.place(x=100,y=180)
빈소기간2.place(x=190,y=180)
안치기간2.place(x=280,y=180)

수납금액.place(x=720,y=10)
받은금액.place(x=720,y=60)
거스름돈.place(x=720,y=110)

#버튼 위치
저장.place(x= 500, y=100)
현금수납.place(x=900, y=10)
닫기.place(x=900, y=70)
식당판매.place(x=700, y=150)
매점판매.place(x=800, y=150)
Set.place(x=900, y=150)


win.mainloop() # 창 실행
