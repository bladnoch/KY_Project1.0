import openpyxl
import tkinter.ttk

def show_in_list(): #시트용 리스트에 저장 --목록하고 번호는 저장 안함
    row=[]
    #원본 시트 사용
    for i in range(2,(get_rows()+1)):
        for j in range(2, 8):
            row.append(ws_data.cell(row=i, column=j).value)
        show_oglist.append(row)
        row = []
def get_rows(): #원본 시트의 rows 길이를 구한다(아이템 숫자+첫 목록)
    count=0

    for rows in ws_data.iter_rows():
        count+=1
    return count

home = "/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/test.xlsx" #기본 물품 엑셀 위치 저장
wb_data = openpyxl.load_workbook(home, data_only=True) #값으로
ws_data=wb_data['Sheet1'] #사용 시트 지정


window=tkinter.Tk()
# window.geometry("640x400+100+100")
# window.resizable(False,False)
show_oglist=[]
show_in_list()

treeview = tkinter.ttk.Treeview(window, columns=["one", "two","three","four","five","six","seven"],
                                displaycolumns=["one", "two","three","four","five","six","seven"])
treeview.pack()

treeview.column("#0", width=30)
treeview.heading("#0", text="번호")

treeview.column("one", width=100, anchor="center")
treeview.heading("one", text="물품코드", anchor="e")

treeview.column("#2", width=100, anchor="w")
treeview.heading("two", text="물품명")

treeview.column("#3", width=50, anchor="w")
treeview.heading("three", text="단위", anchor="e")

treeview.column("#4", width=100, anchor="w")
treeview.heading("four", text="단가", anchor="e")

treeview.column("#5", width=50, anchor="w")
treeview.heading("five", text="수량", anchor="e")

treeview.column("#6", width=50, anchor="w")
treeview.heading("six", text="금액", anchor="center")

treelist=show_oglist


#     [
#     (oglist[1][1],oglist[1][2],oglist[1][3],oglist[1][4],oglist[1][5],oglist[1][6]),
#     (oglist[2][1],oglist[2][2],oglist[2][3],oglist[2][4],oglist[2][5],oglist[2][6])
 # ]

for i in range(len(treelist)):
    treeview.insert('', 'end', text=i+2, values=treelist[i], iid=str(i) + "번")

# top = treeview.insert('', 'end', text=str(len(treelist)), iid="5번", tags="tag1")

window.mainloop()