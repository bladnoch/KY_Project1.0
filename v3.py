from tkinter import * # tkinter의 모든 함수 가져오기
from tkinter import messagebox, filedialog
import os
from pathlib import Path
import openpyxl
import os.path
from openpyxl.worksheet.table import Table, TableStyleInfo
import tkinter.ttk
import tkinter as tk


def del_t(): #오른쪽 트리 삭제용
    tree.delete(*tree.get_children())
def openxl(): #물품추가 버튼 (엑셀 자동으로 열리게 만들면 사라질 예정)
    # os.system(home)
    def close(): #취소 버튼(openxl 종료)
        openxl.quit()
        openxl.destroy()
    def add_in(): # 저장 버튼
        row=[]
        count=0
        for rows in ws.iter_rows():  # ws시트 row 길이를 count에 저장
            count += 1

        for j in range(1,8): #ws시트의 3(물품명),4(단위),5(단가)번 column만 저장하고 6,7 column은 빈 텍스트로 받는다.
            if (j==3): #row에 추가
                row.append(e_item.get())
                ws.cell(row=count+1,column=j).value=e_item.get()
            elif (j==4):
                row.append(e_count.get())
                ws.cell(row=count+1, column=j).value = e_count.get()
            elif(j==5):
                row.append(e_price.get())
                ws.cell(row=count+1, column=j).value = e_price.get()
            else:
                row.append(" ")
                ws.cell(row=count+1, column=j).value = ""
            og_l.append(row) #og_l에 row추가
        wb.save(home) #
        리스트.delete(0,count+1)


        first() #메소드 사용
        close() #창 닫기

    openxl=Tk() #새로운 창 관련

    openxl.geometry("300x170")  # 창의 크기
    openxl.title("물품 추가")  # 창의 제목
    openxl.option_add("*Font", "맑은고딕 11")  # 전체 폰트

    l_item = Label(openxl)
    l_item.config(text="물품명", width=10, relief="solid")
    l_item.place(x=20,y=20)

    l_price = Label(openxl)
    l_price.config(text="단가", width=10, relief="solid")
    l_price.place(x=20, y=50)

    l_count = Label(openxl)
    l_count.config(text="단위", width=10, relief="solid")
    l_count.place(x=20, y=80)

    e_item = Entry(openxl)
    e_item.config(width=20, relief="solid", borderwidth=2)

    e_price = Entry(openxl)
    e_price.config(width=20, relief="solid", borderwidth=2)

    e_count = Entry(openxl)
    e_count.config(width=20, relief="solid", borderwidth=2)

    e_item.place(x=110,y=20)
    e_price.place(x=110,y=50)
    e_count.place(x=110,y=80)

    save = Button(openxl, text="저장")
    save.config(width=6, height=2,command=add_in)
    save.place(x=60,y=115)

    cancel = Button(openxl, text="취소")
    cancel.config(width=6, height=2,command=close)
    cancel.place(x=150,y=115)

    openxl.mainloop()
def close(): #닫기 버튼
    win.quit()
    win.destroy()
def first(): #첫 시작시 실행, if엑셀 파일 생성, og_l에 목록 삽입, 빈 트리 생성
    row=[]
    count=0

    for i in range(1,7): #room 엑셀이 없으면 엑셀 파일 생성
        path = Path(rooms[i])
        if (str(path.is_file()) == "False"):
            nwb = openpyxl.Workbook()  # 엑셀 생성
            info = nwb.create_sheet("info")  # +sheet 이름 1
            items = nwb.create_sheet("items")
            nwb.save(rooms[i])

    for rows in ws.iter_rows(): #기본 물품의 rows 값
        count += 1

    for i in range(1, (count + 1)):  # og_list에 기본 물품 저장
        for j in range(1, 8):
            row.append(ws.cell(row=i, column=j).value)
        og_l.append(row)
        row = []

    for i in range(1,count):
        리스트.insert(i-1,og_l[i][2])


    global tree
    tree = tkinter.ttk.Treeview(win, columns=["one", "two", "three", "four", "five"],
                                displaycolumns=["one", "two", "three", "four", "five"], height=24) #5개 창 생성

    tree.column("#0", width=10, anchor="center") #1
    tree.heading("#0", text="", anchor="center")

    tree.column("#1", width=100, anchor="center") #2
    tree.heading("#1", text="물품명", anchor="center")

    tree.column("#2", width=100, anchor="center") #3
    tree.heading("#2", text="단위", anchor="center")

    tree.column("#3", width=100, anchor="center") #4
    tree.heading("#3", text="단가", anchor="center")

    tree.column("#4", width=100, anchor="center") #5
    tree.heading("#4", text="수량", anchor="center")

    tree.column("#5", width=100, anchor="center") #6
    tree.heading("#5", text="금액", anchor="center")

    if (opener == True): #처음 열때는 빈 tree로 출력
        get = []
        for i in range(1, count):
            for j in range(2, 7):
                get.append("")
            treelist.append(get)
            get = []
        for i in range(len(treelist)):
            tree.insert('', 'end', text="", values=treelist[i])

    else: #처음 연게 아닐경우
        get = []
        for i in range(1, count): #1~row길이만큼 반복
            for j in range(2, 7): #2~6 반복
                get.append(og_l[i][j]) #og_l에 있는 정보를 get리스트에 저장
            treelist.append(get) #treelist에 저장
            get = [] #리스트 get 비운다
        for i in range(len(treelist)): #treelist 길이만큼 반복
            tree.insert('', 'end', text=str(i + 2), values=treelist[i]) #tree에 treelist 입력

    tree.delete(*tree.get_children()) #출력 후 tree를 비운다(다음 받을 tree를 출력하기 위해)
def call_tree(): #아직 사용 안함 (영향 없음)
    count = 0
    for rows in ws.iter_rows():  # 기본 물품의 rows 값
        count += 1

    global tree
    tree= tkinter.ttk.Treeview(win, columns=["one", "two", "three", "four", "five"],
                                    displaycolumns=["one", "two", "three", "four", "five"], height=24)

    tree.column("#0", width=10, anchor="center")
    tree.heading("#0", text="", anchor="center")

    tree.column("#1", width=100, anchor="center")
    tree.heading("#1", text="물품명", anchor="center")

    tree.column("#2", width=100, anchor="center")
    tree.heading("#2", text="단위", anchor="center")

    tree.column("#3", width=100, anchor="center")
    tree.heading("#3", text="단가", anchor="center")

    tree.column("#4", width=100, anchor="center")
    tree.heading("#4", text="수량", anchor="center")

    tree.column("#5", width=100, anchor="center")
    tree.heading("#5", text="금액", anchor="center")


    get = []
    for i in range(1, count):
        for j in range(2, 7):
            get.append(og_l[i][j])
        treelist.append(get)
        get = []
    for i in range(len(treelist)):
        tree.insert('', 'end', text=str(i + 2), values=treelist[i])

    tree.delete(*tree.get_children())
    win.update()
def check(): #값 출력해서 확인하는 용도 (영향 없음)
    count = 0
    for rows in ws.iter_rows():  # 기본 물품의 rows 값
        count += 1
    path = Path(room1)
    messagebox.showinfo("", str(path.is_file()))
    # messagebox.showinfo("",)
def save(): #저장 버튼 (저장관련: 빈소, 고인명, 상주명, id 확인)
    room=빈소.get()
    pinfo=[room,고인명.get(),상주명.get(),ID.get()] #개인정보 pinfo리스트에 추가
    empty=False

    for i in range (len(pinfo)): #pinfo 리스트를 반복해서 리스트에 정보가 하나라도 없으면 empty=True
        if(pinfo[i]==""):
            empty=True
    if(empty==True): #empty가 True면 메세지 박스 실행
        messagebox.showinfo("", "정보를 입력해 주세요")
        empty=False #다시 버튼이 눌렸을때를 위해 다시 False로 바꿔준다
    else: #모든 정보가 있어서 empty가 False면 save_go() 실행
        save_go()

    # messagebox.showinfo("","빈소"+room+"에 저장 하시겠습니까?")
def save_go(): #실제 저장 실행
    room = 빈소.get() #빈소 호수 저장
    if ((room=="1")|(room=="2")|(room=="3")|(room=="4")|(room=="5")|(room=="6")): #room이 1~6사이이면 실행

        # 빈소에 넣은 숫자에 따라 사용하는 엑셀이 달라짐
        if (room == "1"):
            nwb = openpyxl.load_workbook(room1) #room1엑셀을 불러온다
            info = nwb["info"]  #sheet 이름 1
            # items = nwb["items"]  #sheet 이름 2
            nwb.remove(nwb["items"]) #items 시트 삭제. 덮어쓸때 기존 시트 정보와 새로운 정보가 같이 나오기 때문에
            items = nwb.create_sheet("items") #items 시트 다시 생성.

            #개인정보 저장 위치
            info['A1'] = ID.get()
            info['B1'] = 고인명.get()
            info['C1'] = 상주명.get()
            info['D1'] = room

            for i in range(len(new_l)): #트리에 있던 값 저장(new_l)의 길이만큼 반복
                for j in range(5): #필요한 정보 5개만큼 반복
                    # loc=alp[j]+str(i)
                    items.cell(row=i+1,column=j+1).value=new_l[i][j] #items 시트에 new_l의 정보 저장
                    # messagebox.showinfo("",new_l[i][j])

            nwb.save(room1) #엑셀 파일 저장

        elif (room == "2"):
            nwb = openpyxl.load_workbook(room2)
            info = nwb["info"]  # +sheet 이름 1
            # items = nwb["items"]  # +sheet 이름 2
            nwb.remove(nwb["items"])
            items = nwb.create_sheet("items")

            info['A1'] = ID.get()
            info['B1'] = 고인명.get()
            info['C1'] = 상주명.get()
            info['D1'] = room


            for i in range(len(new_l)): #트리에 있던 값 저장(new_l)
                for j in range(5):
                    # loc=alp[j]+str(i)
                    items.cell(row=i+1,column=j+1).value=new_l[i][j]
                    # messagebox.showinfo("",new_l[i][j])

            nwb.save(room2)

        elif (room == "3"):
            nwb = openpyxl.load_workbook(room3)
            info = nwb["info"]  # +sheet 이름 1
            # items = nwb["items"]  # +sheet 이름 2
            nwb.remove(nwb["items"])
            items = nwb.create_sheet("items")

            if (ID.get() == ""):
                info['A1'] = " "
            if (상주명.get() == ""):
                info['A1'] = " "
            if (고인명.get() == ""):
                info['A1'] = " "

            info['A1'] = ID.get()
            info['B1'] = 고인명.get()
            info['C1'] = 상주명.get()
            info['D1'] = room

            for i in range(len(new_l)): #트리에 있던 값 저장(new_l)
                for j in range(5):
                    # loc=alp[j]+str(i)
                    items.cell(row=i+1,column=j+1).value=new_l[i][j]
                    # messagebox.showinfo("",new_l[i][j])

            nwb.save(room3)

        elif (room == "4"):
            nwb = openpyxl.load_workbook(room4)
            info = nwb["info"]  # +sheet 이름 1
            # items = nwb["items"]  # +sheet 이름 2
            nwb.remove(nwb["items"])
            items = nwb.create_sheet("items")

            info['A1'] = ID.get()
            info['B1'] = 고인명.get()
            info['C1'] = 상주명.get()
            info['D1'] = room

            for i in range(len(new_l)): #트리에 있던 값 저장(new_l)
                for j in range(5):
                    # loc=alp[j]+str(i)
                    items.cell(row=i+1,column=j+1).value=new_l[i][j]
                    # messagebox.showinfo("",new_l[i][j])

            nwb.save(room4)

        elif (room == "5"):
            nwb = openpyxl.load_workbook(room5)
            info = nwb["info"]  # +sheet 이름 1
            # items = nwb["items"]  # +sheet 이름 2
            nwb.remove(nwb["items"])
            items = nwb.create_sheet("items")

            info['A1'] = ID.get()
            info['B1'] = 고인명.get()
            info['C1'] = 상주명.get()
            info['D1'] = room

            for i in range(len(new_l)): #트리에 있던 값 저장(new_l)
                for j in range(5):
                    # loc=alp[j]+str(i)
                    items.cell(row=i+1,column=j+1).value=new_l[i][j]
                    # messagebox.showinfo("",new_l[i][j])

            nwb.save(room5)

        elif (room == "6"):
            nwb = openpyxl.load_workbook(room6)
            info = nwb["info"]  # +sheet 이름 1
            # items = nwb["items"]  # +sheet 이름 2
            nwb.remove(nwb["items"])
            items = nwb.create_sheet("items")

            info['A1'] = ID.get()
            info['B1'] = 고인명.get()
            info['C1'] = 상주명.get()
            info['D1'] = room

            for i in range(len(new_l)): #트리에 있던 값 저장(new_l)
                for j in range(5):
                    # loc=alp[j]+str(i)
                    items.cell(row=i+1,column=j+1).value=new_l[i][j]
                    # messagebox.showinfo("",new_l[i][j])

            nwb.save(room6)
    else:
        messagebox.showinfo('없는 빈소',"정확한 빈소명을 입력해주세요")
def clickEvent(event): #리스트박스 더블 클릭하면 실행(인덱스 받아서 tree에 추가)
    eventNum=list(리스트.curselection()) #누른 물건의 인덱스를 받는다
    num=eventNum[0] #받은 인덱스의 형을 인트로 받는다
    # messagebox.showinfo("",eventNum)

    row=[]
    count=0
    for rows in ws.iter_rows(): #기본 물품의 rows 길이 저장
        count += 1

    for i in range(1, (count + 1)):  # og_list에 기본 물품 저장
        for j in range(1, 8):
            row.append(ws.cell(row=i, column=j).value)
        og_l.append(row) #og_l에 왼쪽 목록의 물건들과 정보 저장
        row = []

    # messagebox.showinfo("",str(og_l[num+1][2])+" "+str(og_l[num+1][3])+" "+str(og_l[num+1][4])+" "+str(og_l[num+1][5]))


    global tree #전역변수 tree 생성
    del tree #뭐더라...
    treelist = []

    tree= tkinter.ttk.Treeview(win, columns=["one", "two", "three", "four", "five"],
                                    displaycolumns=["one", "two", "three", "four", "five"], height=24) #tree생성 column 5개

    tree.column("#0", width=10, anchor="center")
    tree.heading("#0", text="", anchor="center")

    tree.column("#1", width=100, anchor="center")
    tree.heading("#1", text="물품명", anchor="center")

    tree.column("#2", width=100, anchor="center")
    tree.heading("#2", text="단위", anchor="center")

    tree.column("#3", width=100, anchor="center")
    tree.heading("#3", text="단가", anchor="center")

    tree.column("#4", width=100, anchor="center")
    tree.heading("#4", text="수량", anchor="center")

    tree.column("#5", width=100, anchor="center")
    tree.heading("#5", text="금액", anchor="center")


    get = []

    for i in range(2,7):#정보 입력을 위해 og_l과 num(index 위치)를 이용해 new_l(오른쪽 목록에 추가되는 물건들의 정보를 저장)에 저장
        if(i==5): #5는 수량 수량지정을 안 만들어서 무조건 1로 고정
            og_l[num+1][i]=1
        elif(i==6): #(4)단가* (5)수량을 6변째 column에 저장
            og_l[num+1][i]=og_l[num+1][i-2]*og_l[num+1][i-1]
        get.append(og_l[num+1][i])
    # treelist.append(get)
    new_l.append(get) #최종적으로 new_l에 저장

    del_t()
        # messagebox.showinfo("", og_l[num+1][i])
    if (len(new_l)>=1):
        for i in range(len(new_l)): #new_l의 정보를 tree에 추가

            tree.insert('', 'end', text="", values=new_l[i])
            # messagebox.showinfo("", new_l[i])
            # messagebox.showinfo("", len(new_l))
        # messagebox.showinfo("treelist[i]",new_l[i])

        tree.place(x=170,y=210) #새로운 tree 위치 지정
        tree.bind("<Double-Button-1>", clickEvent_delete) #이 새로생긴 트리에 함수를 지정(tree를 더블 클릭하면 물건을 지우기 위해)

    # get.clear()
    # tree.delete(*tree.get_children())
    win.update() #win화면 업데이트
def clickEvent_delete(event): #tree 더블클릭하면 실행
    selectedItem=tree.selection()[0] #tree 선택한 위치 받기
    # messagebox.showinfo("",tree.item(selectedItem)['values'][0])
    # messagebox.showinfo("",len(new_l))
    for i in range(len(new_l)): #삭제될 tree 요소를 list에서도 삭제
        if(tree.item(selectedItem)['values'][0]==new_l[i][0]): #선택된 트리의 값과 new_l에 저장되있던 값이 겹치는게 있으면
            new_l.remove(new_l[i]) #겹쳐진 new_l의 row 값을 삭제
            break;

    # messagebox.showinfo("",tree.item(selectedItem)['values'][1])

    selected_item = tree.selection()[0]  ## get selected item
    # new_l=[]
    tree.delete(selected_item) #트리에 저장된 정보 삭제
def loadxl(): #불러오기 버튼
    def close(): #제삭
        loadxl.quit()
        loadxl.destroy()
    def load_btn():
        def click_delete(event): #트리에서 항목 더블클리(삭제)
            selectedItem = tree.selection()[0]
            for i in range(len(new_l)):  # 선택한 트리속 목록을 new_l과 비교
                if (tree.item(selectedItem)['values'][0] == new_l[i][0]): #new_l을 확인하고
                    new_l.remove(new_l[i]) #new_l에서 삭제
                    break;
            selection = tree.selection()[0]
            tree.delete(selection) #선택된 항목 트리에서 삭제

        roomNum=l_room.get() #입력한 호수 저장
        # messagebox.showinfo("",roomNum)

        rooms=[room1,room2,room3,room4,room5,room6] #리스트 rooms에 방 6개의 주소 저장

        for i in range(6): #사용할 시트를 찾아서 시트에 맞는 info,items 사용
            nwb = openpyxl.load_workbook(rooms[i])
            info = nwb["info"]
            items = nwb["items"]

            if (roomNum==info.cell(row=1, column=4).value): #맞는 빈소를 찾으면 멈추고 for문 종료
                # messagebox.showinfo("", "True")
                break

        ID.delete(0,END)
        고인명.delete(0, END)
        상주명.delete(0, END)
        빈소.delete(0, END)

        ID.insert(0,info.cell(row=1, column=1).value)
        고인명.insert(0,info.cell(row=1, column=2).value)
        상주명.insert(0,info.cell(row=1, column=3).value)
        빈소.insert(0,info.cell(row=1, column=4).value)


        tree = tkinter.ttk.Treeview(win, columns=["one", "two", "three", "four", "five"],
                                    displaycolumns=["one", "two", "three", "four", "five"], height=24)

        tree.column("#0", width=10, anchor="center")
        tree.heading("#0", text="", anchor="center")

        tree.column("#1", width=100, anchor="center")
        tree.heading("#1", text="물품명", anchor="center")

        tree.column("#2", width=100, anchor="center")
        tree.heading("#2", text="단위", anchor="center")

        tree.column("#3", width=100, anchor="center")
        tree.heading("#3", text="단가", anchor="center")

        tree.column("#4", width=100, anchor="center")
        tree.heading("#4", text="수량", anchor="center")

        tree.column("#5", width=100, anchor="center")
        tree.heading("#5", text="금액", anchor="center")



        count=0
        for rows in items.iter_rows():  # 선택된 파일의 items시트의 rows 값
            count += 1
        # messagebox.showinfo("","선택한 시트의 rows 갯수"+str(count))


        put=[]
        get=[]
        new_l.clear()
        for i in range(1,count+1):
            for j in range(1,6):
                get.append(items.cell(row=i,column=j).value)
                # messagebox.showinfo("",items.cell(row=i,column=j).value)
            put.append(get)
            new_l.append(get) #new_l에 저장
            get=[]
        # messagebox.showinfo("",len(put))
        for i in range(len(put)):
            tree.insert('', 'end', text=" ", values=put[i]) #tree에 출력


        tree.place(x=170, y=210)
        tree.bind("<Double-Button-1>", click_delete) #트리 더블클릭하면 실행


        # selected_item = tree.item()  ## get selected item
        # tree.delete(selected_item)
        # clear_new_l()
        # messagebox.showinfo(new_l[0])
        # tree.bind("<Double-Button-1>", click_del)

        close()



    loadxl=Tk() #불러오기 하면 나오는 화면

    loadxl.geometry("300x120")  # 창의 크기
    loadxl.title("불러오기")  # 창의 제목
    loadxl.option_add("*Font", "맑은고딕 11")  # 전체 폰트

    l_room = Label(loadxl)
    l_room.config(text="빈소 호수", width=10, relief="solid")
    l_room.place(x=20,y=20)

    l_room = Entry(loadxl)
    l_room.config(width=20, relief="solid", borderwidth=2)

    l_room.place(x=110,y=20)

    load = Button(loadxl, text="저장")
    load.config(width=10, height=3,command=load_btn)
    load.place(x=35,y=55)

    cancel = Button(loadxl, text="취소")
    cancel.config(width=10, height=3,command=close)
    cancel.place(x=145,y=55)

    loadxl.mainloop()
def openfile(): #작동 안됨 영향 없음
    os.system(home)
def shifter1(): #시연용 급하게 추가한 함수
    리스트.delete(0,END)
    og_l = []
    count=0
    row=[]
    for rows in ws.iter_rows():  # 기본 물품의 rows 값
        count += 1

    for i in range(1, (count + 1)):  # og_list에 기본 물품 저장
        for j in range(1, 8):
            row.append(ws.cell(row=i, column=j).value)
        og_l.append(row)
        row = []

    for i in range(1, count):
        리스트.insert(i - 1, og_l[i][2])
def shifter2(): #시연용 급하게 추가한 함수
    리스트.delete(0,END)
    og_l = []
    count=0
    for rows in ws2.iter_rows():  # 기본 물품의 rows 값
        count += 1
    row=[]
    for i in range(1, (count + 1)):  # og_list에 기본 물품 저장
        for j in range(1, 8):
            row.append(ws2.cell(row=i, column=j).value)
        og_l.append(row)
        row = []

    for i in range(1, count):
        리스트.insert(i - 1, og_l[i][2])
def shifter3(): #시연용 급하게 추가한 함수
    리스트.delete(0,END)
    og_l=[]
    # messagebox.showinfo("","hello")
    count=0
    row=[]
    for rows in ws3.iter_rows():  # 기본 물품의 rows 값
        count += 1
    # messagebox.showinfo("","world")
    for i in range(1, (count + 1)):  # og_list에 기본 물품 저장
        for j in range(1, 8):
            row.append(ws3.cell(row=i, column=j).value)
        og_l.append(row)
        row = []

    for i in range(1, count):
        리스트.insert(i - 1, og_l[i][2])
        # messagebox.showinfo("",리스트)
    win.update()
##################################################   global variable   ##########################

home = 'excelhere/test.xlsx' #기본 물품 엑셀 위치 저장

room1='excelhere/room_one.xlsx'
room2='excelhere/room_two.xlsx'
room3='excelhere/room_three.xlsx'
room4='excelhere/room_four.xlsx'
room5='excelhere/room_five.xlsx'
room6='excelhere/room_six.xlsx'

nwb = openpyxl.Workbook()  # 엑셀 생성
info=nwb.create_sheet("info")  # +sheet 이름 1
items=nwb.create_sheet("items")  # +sheet 이름 2

wb= openpyxl.load_workbook(home, data_only=True) #초기 시트 위치 저장(값으로)
ws=wb['Sheet1'] #초기 시트 사용 선언
ws2=wb['Sheet2']
ws3=wb['Sheet3']
wslist=[ws,ws2,ws3]


global rooms
rooms=['',room1,room2,room3,room4,room5,room6]
global og_l #초기 리스트 저장공간
og_l=[]
global new_l #새로운 리스트 저장공간
new_l=[]
global treelist #list
treelist=[]
global opener
opener=True
global c_table
c_table=False



##################################################   tkinter   ##########################
win = tk.Tk() # 창 생성
win.geometry("1200x720") # 창의 크기
win.title("장례식장 재고관리 프로그램 Ver1.221123") # 창의 제목
win.option_add("*Font", "맑은고딕 12") # 전체 폰트


#win.resizable(False, False) #윈도우 사이즈 조절 불가

ID_lab = Label(win)
ID_lab.config(text = "ID", width=10, relief="solid")
물품명 = Label(win)
물품명.config(text = "물품명", width=15, relief="solid",borderwidth=0)
고인명_lab = Label(win)
고인명_lab.config(text = "고인명",width=10, relief="solid")
상주명_lab = Label(win)
상주명_lab.config(text = "상주명",width=10, relief="solid")
빈소_lab = Label(win)
빈소_lab.config(text = "빈소", width=10, relief="solid")

기간 = Label(win)
기간.config(text = "빈소 기간 : XXXX / XX / XX ~ XXXX / XX / XX \n안치 기간 : XXXX / XX / XX ~ XXXX / XX / XX", width=47, relief="solid",height=3)

수납금액_lab = Label(win)
수납금액_lab.config(text = "수납금액", width=10, relief="solid")
받은금액_lab = Label(win)
받은금액_lab.config(text = "받은금액", width=10, relief="solid")
거스름돈_lab = Label(win)
거스름돈_lab.config(text = "거스름돈", width=10, relief="solid")


##################################################   entry   ##########################
ID = Entry(win)
ID.config(width=10,relief="solid",borderwidth=2)
고인명 = Entry(win)
고인명.config(width=10,relief="solid",borderwidth=2)
상주명 = Entry(win)
상주명.config(width=10,relief="solid",borderwidth=2)
빈소 = Entry(win)
빈소.config(width=10,relief="solid",borderwidth=2)

수납금액 = Entry(win)
수납금액.config(width=20,relief="solid",borderwidth=2)
받은금액 = Entry(win)
받은금액.config(width=20,relief="solid",borderwidth=2)
거스름돈 = Entry(win)
거스름돈.config(width=20,relief="solid",borderwidth=2)
리스트 = Listbox(win, selectmode = 'extended',width = 15, height = 27,borderwidth=0)
리스트.bind("<Double-Button-1>", clickEvent)
리스트.yview()

##################################################   buttons   ##########################
저장 = Button(win, text = "저장")
저장.config(width=14,height=2,command=save)
불러오기 = Button(win, text = "불러오기")
불러오기.config(width=14,height=2, command=loadxl)
닫기 = Button(win, text = "닫기",command=close)
닫기.config(width=10,height=3)
물품추가 = Button(win, text = "물품추가",command=openxl)
물품추가.config(width=10,height=1)
물품삭제 = Button(win, text = "물품삭제")
물품삭제.config(width=7,height=2)

시트1= Button(win, text = "시트1")
시트1.config(width=7,height=2,command=shifter1)
시트2 = Button(win, text = "시트2")
시트2.config(width=7,height=2,command=shifter2)
시트3 = Button(win, text = "시트3")
시트3.config(width=7,height=2,command=shifter3)

물품비우기 = Button(win, text = "물품 비우기")
물품비우기.config(width=7,height=2)
Set = Button(win, text = "프린트")
Set.config(width=10,height=3, command=openfile)

first()


##################################################   place   ##########################
#labels
ID_lab.place(x=10,y=10)
고인명_lab.place(x=210,y=10)
상주명_lab.place(x=210,y=50)
빈소_lab.place(x=10,y=50)
수납금액_lab.place(x=720,y=10)
받은금액_lab.place(x=720,y=50)
거스름돈_lab.place(x=720,y=90)
기간.place(x=10,y=90)

#엔트리 위치
ID.place(x=100,y=10)
고인명.place(x=300,y=10)
상주명.place(x=300,y=50)
빈소.place(x=100,y=50)


수납금액.place(x=810,y=10)
받은금액.place(x=810,y=50)
거스름돈.place(x=810,y=90)

#버튼 위치

시트1.place(x=10,y=100)
시트2.place(x=10,y=140)
시트3.place(x=10,y=180)

저장.place(x= 310, y=150)
불러오기.place(x=460,y=150)
닫기.place(x=860, y=120)
물품추가.place(x=148, y=157)
# 물품삭제.place(x=700, y=310)
# 물품비우기.place(x=700, y=230)
Set.place(x=730, y=120)
tree.place(x=170,y=210)
리스트.place(x=48, y=236)


win.mainloop() # 창 실행