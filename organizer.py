from tkinter import * # tkinter의 모든 함수 가져오기
from tkinter import messagebox
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import tkinter.ttk

def close():
    win.quit()
    win.destroy()
def set_xl(): #시작하면 바로 시작
    row = []
    count=0
    global treelist
    treelist=[]
    for rows in ws.iter_rows(): #기본 물품의 rows 값
        count += 1
    # 원본 시트 사용
    for i in range(1, (count + 1)): #og_list에 기본 물품 저장
        for j in range(1, 8):
            row.append(ws.cell(row=i, column=j).value)
        og_list.append(row)
        row = []

    count = 0
    for rows in ws.iter_rows():
        count += 1

    global tree
    tree = tkinter.ttk.Treeview(win, columns=["one", "two", "three", "four", "five", "six"],
                                displaycolumns=["one", "two", "three", "four", "five", "six"], height=25)

    tree.column("#0", width=40, anchor="center")
    tree.heading("#0", text="번호", anchor="center")

    tree.column("#1", width=100, anchor="center")
    tree.heading("#1", text="물품코드", anchor="center")

    tree.column("#2", width=100, anchor="center")
    tree.heading("#2", text="물품명", anchor="center")

    tree.column("#3", width=100, anchor="center")
    tree.heading("#3", text="단위", anchor="center")

    tree.column("#4", width=100, anchor="center")
    tree.heading("#4", text="단가", anchor="center")

    tree.column("#5", width=100, anchor="center")
    tree.heading("#5", text="수량", anchor="center")

    tree.column("#6", width=100, anchor="center")
    tree.heading("#6", text="금액", anchor="center")

    get = []
    for i in range(1, count):
        for j in range(1, 7):
            get.append(og_list[i][j])
        treelist.append(get)
        get = []
    for i in range(len(treelist)):
        tree.insert('', 'end', text=i + 2, values=treelist[i])
def save(): #저장 눌렀을때 작동. 개인정보, 시트 지정해서 저장
    if (빈소.get()==""):
        messagebox.showinfo("","빈소를 정해주세요")
    else:
        room = 빈소.get()


        info['A1'] = ID.get()
        info['B1'] = 고인명.get()
        info['C1'] = 상주명.get()
        info['D1'] = 빈소.get()

        # 빈소에 넣은 숫자에 따라 사용하는 엑셀이 달라짐
        if (room == "1"):
            nwb.save(room1)
        elif (room == "2"):
            nwb.save(room2)
        elif (room == "3"):
            nwb.save(room3)
        elif (room == "4"):
            nwb.save(room4)
        elif (room == "5"):
            nwb.save(room5)
        elif (room == "6"):
            nwb.save(room6)

def checker():
    messagebox.showinfo("빈소.get()",visiter)

def get_rows():
    if (visiter==0):
        count = 0
        for rows in ws.iter_rows():
            count += 1
        return count
def get_rows(sheet):
    count = 0
    for rows in sheet.iter_rows():
        count += 1
    return count
def call_tree():
    global tree
    tree= tkinter.ttk.Treeview(win, columns=["one", "two", "three", "four", "five", "six"],
                                    displaycolumns=["one", "two", "three", "four", "five", "six"], height=25)

    tree.column("#0", width=40, anchor="center")
    tree.heading("#0", text="번호", anchor="center")

    tree.column("#1", width=100, anchor="center")
    tree.heading("#1", text="물품코드", anchor="center")

    tree.column("#2", width=100, anchor="center")
    tree.heading("#2", text="물품명", anchor="center")

    tree.column("#3", width=100, anchor="center")
    tree.heading("#3", text="단위", anchor="center")

    tree.column("#4", width=100, anchor="center")
    tree.heading("#4", text="단가", anchor="center")

    tree.column("#5", width=100, anchor="center")
    tree.heading("#5", text="수량", anchor="center")

    tree.column("#6", width=100, anchor="center")
    tree.heading("#6", text="금액", anchor="center")

    get = []
    for i in range(1, get_rows()):
        for j in range(1, 7):
            get.append(og_list[i][j])
        treelist.append(get)
        get = []
    for i in range(len(treelist)):
        tree.insert('', 'end', text=i + 2, values=treelist[i])


##################################################   global variable   ##########################
home = "/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/test.xlsx" #기본 물품 엑셀 위치 저장

room1='/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/room_one.xlsx'
room2='/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/room_two.xlsx'
room3='/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/room_three.xlsx'
room4='/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/room_four.xlsx'
room5='/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/room_five.xlsx'
room6='/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/room_six.xlsx'

nwb = openpyxl.Workbook()  # 엑셀 생성
info=nwb.create_sheet("info")  # +sheet 이름 1
items=nwb.create_sheet("items")  # +sheet 이름 2

visiter="0" #사람 정보용
og_list=[] #기본 물품 엑셀
new_list=[] #새 물품 엑셀
wb= openpyxl.load_workbook(home, data_only=True) #값으로
ws=wb['Sheet1'] #사용 시트 지정




##################################################   tkinter   ##########################
win = Tk() # 창 생성
win.geometry("1000x720") # 창의 크기
win.title("장례식장 재고관리 프로그램 Ver1.221123") # 창의 제목
win.option_add("*Font", "맑은고딕 11") # 전체 폰트
#win.resizable(False, False) #윈도우 사이즈 조절 불가


##################################################   labels   ##########################
ID_lab = Label(win)
ID_lab.config(text = "ID", width=10, relief="solid")
고인명_lab = Label(win)
고인명_lab.config(text = "고인명",width=10, relief="solid")
상주명_lab = Label(win)
상주명_lab.config(text = "상주명",width=10, relief="solid")
빈소_lab = Label(win)
빈소_lab.config(text = "빈소", width=10, relief="solid")

수납금액_lab = Label(win)
수납금액_lab.config(text = "수납금액", width=10, relief="solid")
받은금액_lab = Label(win)
받은금액_lab.config(text = "받은금액", width=10, relief="solid")
거스름돈_lab = Label(win)
거스름돈_lab.config(text = "거스름돈", width=10, relief="solid")


##################################################   entry   ##########################
ID = Entry(win)
ID.config(width=10,relief="solid",borderwidth=2)
고인명 = Entry(win)
고인명.config(width=10,relief="solid",borderwidth=2)
상주명 = Entry(win)
상주명.config(width=10,relief="solid",borderwidth=2)
빈소 = Entry(win)
빈소.config(width=30,relief="solid",borderwidth=2)

수납금액 = Entry(win)
수납금액.config(width=20,relief="solid",borderwidth=2)
받은금액 = Entry(win)
받은금액.config(width=20,relief="solid",borderwidth=2)
거스름돈 = Entry(win)
거스름돈.config(width=20,relief="solid",borderwidth=2)


##################################################   buttons   ##########################
저장 = Button(win, text = "저장")
저장.config(width=10,height=2, command=save)
불러오기 = Button(win, text = "불러오기")
불러오기.config(width=10,height=2)
닫기 = Button(win, text = "닫기")
닫기.config(width=10,height=3,command =close)
수정 = Button(win, text = "수정")
수정.config(width=10,height=3)
삭제 = Button(win, text = "삭제")
삭제.config(width=10,height=3)
Set = Button(win, text = "checker")
Set.config(width=10,height=3, command=checker)


set_xl()
# call_tree()
##################################################   place   ##########################
#labels
ID_lab.place(x=10,y=10)
고인명_lab.place(x=210,y=10)
상주명_lab.place(x=410,y=10)
빈소_lab.place(x=10,y=50)
수납금액_lab.place(x=620,y=10)
받은금액_lab.place(x=620,y=60)
거스름돈_lab.place(x=620,y=110)

#엔트리 위치
ID.place(x=110,y=10)
고인명.place(x=310,y=10)
상주명.place(x=510,y=10)
빈소.place(x=110,y=50)


수납금액.place(x=720,y=10)
받은금액.place(x=720,y=60)
거스름돈.place(x=720,y=110)

#버튼 위치
tree.place(x=10,y=210)
저장.place(x= 350, y=50)
불러오기.place(x=470,y=50)
닫기.place(x=370, y=150)
수정.place(x=10, y=150)
삭제.place(x=130, y=150)
Set.place(x=250, y=150)

win.mainloop() # 창 실행