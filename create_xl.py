import openpyxl


# # 엑셀 파일 만들기
# wb=openpyxl.Workbook()
#
# #엑셀 워크시트 만들기
# ws=wb.create_sheet("워크시트 만들기")

#데이터 추가하기
#test.xlsx를 불러 올수 있으면 불러와서 이곳에 추가
#자동으로 위치를 지정하는 방법이 있어야함

nwb=openpyxl.Workbook()
pws=nwb.create_sheet("private_info")
iws=nwb.create_sheet("items")

pws['A1']='1'
pws['B1']='물품코드'
pws['C1']='물품명'

iws['A1']=2
iws['B1']="0001" #물품명 string
iws['C1']='고무장갑'

nwb.save('/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/personal.xlsx')

# test="A"
# test2=1
# for i in range:
#     test2=i
#     test3=test+test2
#     print(test3)
# ws['A1']='1'
# ws['B1']='물품코드'
# ws['C1']='물품명'
# ws['D1']='단위'
# ws['E1']='단가'
# ws['F1']='수량'
# ws['G1']='금액'
#
# ws['A1']=2
# ws['B1']="0001" #물품명 string
# ws['C1']='고무장갑'
# ws['D1']='개'
# ws['E1']=2300
# ws['F1']=''
# ws['G1']=0


#엑셀 저장
# !엑셀 이름을 어떻게 줄것인지 생각해야함 ->int변수나 id를 이름으로 주는 방법
nwb.save('/Users/doungukkim/Desktop/workspace/python/restinpeace/excelhere/personal.xlsx')